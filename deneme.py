from requests.auth import HTTPBasicAuth
from requests.auth import HTTPDigestAuth
import requests
from hikvisionapi import Client 
from openpyxl import Workbook,load_workbook
import sys, os, re
from requests import Session
import urllib3
import time

init()
ip_list=[]
cam=''
value_table=[]
user=''
password=''
stream_value_table=[]
qos_value_table=[]
YeniData=[]
privacyMaskEnable=''
PutorGet =''
m=''
print_value=''

def camera_connection(ip,user,password):
    try:
        connected_camera=Client('http://'+ip,user,password)
        return connected_camera
    except:
        print("XML bağlantı yolunu kontrol et!")
        return 0

def XMLformat(cam):
    try:
        if (print_value=='speed') or (print_value=='duplex') or (print_value=='MTU'):
            XMLL=cam.System.Network.interfaces['1'].link(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='upnp'):
            XMLL=cam.System.Network.interfaces['1'].discovery(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='qos'):
            XMLL=cam.System.Network.qos.dscp(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='multicast') or (print_value=='multicast2') or (print_value=='multicast3') or (print_value=='mainstream') or (print_value=='secondstream') or (print_value=='thirdstream') :
            XMLL = cam.Streaming.channels[m](method=PutorGet,present='text',data=YeniData)
        elif (print_value=='deviceName') or (print_value=='version'):
            XMLL = cam.System.deviceInfo(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='OnlineUser'):
            XMLL = cam.Security.OnlineUser(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='ntp'):
            XMLL = cam.System.time.ntpservers(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='motionDetection'):
            XMLL = cam.System.Video.inputs.channels['1'].motionDetection(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='privacyMask'):
            XMLL = cam.System.video.inputs.channels['1'].privacyMask(method=PutorGet,present='text',data=YeniData)       
        elif (print_value=='ssh'):
            XMLL = cam.System.Network.ssh(method=PutorGet,present='text',data=YeniData)
        elif (print_value=='IEEE802_1x'):
            print("xxx")
            XMLL = cam.System.Network.IEEE802_1x(method=PutorGet,present='text',data=YeniData)
            print(XMLL)
        else:
            print("")
        return XMLL
    except:        
        print("XML bağlantı yolunu kontrol et")
        return 0
#*********************************************************************************************************************************
counter_for_null_value=0
PutorGet=str(input("'put' or 'get':").strip()) #print ettirmek istediğiniz toplu değerler  
if (PutorGet.lower() =='get' or PutorGet.lower() == "put"):    
    print("Tüm degerleri excele yazdirmak icin 'all' yaziniz: \n ya da 'deviceName','multicast','multicast2','multicast3','speed','duplex','qos','version','ntp','MTU','upnp','mainstream','secondstream','thirdstream','motionDetection','OnlineUserNumber','bitrate', 'ssh' Degerlerinden birini yazabilirsiniz\n Yazdirmak istediginiz degerler bittiginde 'end' yazarak islemi sonlandirabilirsiniz.") #print ettirmek istediğiniz toplu değerler
    selected_values=""    
    while counter_for_null_value < 4 and not (selected_values=='end'):
        selected_values=str(input(" Yazdirmak istediginiz değeri giriniz:").strip())
        if(selected_values ==""):
            counter_for_null_value += 1        
        elif (selected_values=='all'):
            value_table=['deviceName','multicast','multicast2','multicast3','speed','duplex','qos','version','ntp','MTU','upnp','privacyMask','mainstream','secondstream','thirdstream','motionDetection','OnlineUserNumber','bitrate','ssh','IEEE802_1x', 'end']
            stream_value_table=['resolution','encoding','bitratetype','maxbitrate','framehizi','iframe','end']
            qos_value_table=['videosesdscp','olayalarmdscp','dscpyonetimi','end']
            counter_for_null_value = 4         
        elif (selected_values.lower() == "devicename"or selected_values.lower() == "ssh" or selected_values.lower() == "multicast" or selected_values.lower() == "multicast2" or selected_values.lower() == "multicast3" or selected_values.lower() == "speed" or selected_values.lower() == "version" or selected_values.lower() == "privacymask" or selected_values.lower() == "duplex" or selected_values.lower() == "qos" or selected_values.lower() == "MTU" or selected_values.lower() == "ntp" or selected_values.lower() == "upnp" or selected_values.lower() == "mainstream" or selected_values.lower() == "secondstream" or selected_values.lower() == "thirdstream" or selected_values.lower() == "motionDetection" or selected_values =='IEEE802_1x' or selected_values.lower() == 'onlineusernumber'):
            if (selected_values.lower()=='privacymask'):
                selected_values="privacyMask"
                privacyMaskEnable=str(input("PrivacyMask True or False:").strip()) #print ettirmek istediğiniz toplu değerler
                privacyMaskEnable=privacyMaskEnable.lower()            
            elif(selected_values=='qos'):
                print("'VideoSesDSCP','OlayAlarmDSCP','DSCPYonetimi','all' Degerlerinden birini yazabilirsiniz")
                selected_values_for_qos=""            
                while not (selected_values_for_qos =='end' or selected_values_for_qos =='all'):                    
                    selected_values_for_qos=str(input("Qos içinde hangi değerleri yazdırmak istersiniz:").strip())
                    selected_values_for_qos=selected_values_for_qos.lower()            
                    if(selected_values_for_qos ==""):
                        counter_for_null_value += 1
                    elif (selected_values_for_qos == "videosesdscp" or selected_values_for_qos == "olayalarmdscp" or selected_values_for_qos == "dscpyonetimi"):
                        qos_value_table.append(selected_values_for_qos)            
                    elif selected_values_for_qos =='all':
                        qos_value_table=['videosesdscp','olayalarmdscp','dscpyonetimi','end']            
                    elif(selected_values_for_qos == "end"):
                         print("qos degerleri alindi.")          
                    else: 
                        print ("Deger tanimlanamadi.")            
            elif (selected_values=='mainstream') or (selected_values=='secondstream') or (selected_values=='thirdstream'):
                print("'resolution','encoding','bitratetype','maxbitrate','framehizi','iframe' Degerlerinden birini yazabilirsiniz")
                selected_values_for_stream=""            
                while not (selected_values_for_stream =='end' or selected_values_for_stream == 'all'):                    
                    selected_values_for_stream=str(input(" Stream içinde hangi değerleri yazdırmak istersiniz:").strip())
                    selected_values_for_stream=selected_values_for_stream.lower()            
                    if(selected_values_for_stream ==""):
                        counter_for_null_value += 1
                    if (selected_values_for_stream == "resolution" or selected_values_for_stream == "encoding" or selected_values_for_stream == "bitratetype" or selected_values_for_stream == "maxbitrate" or selected_values_for_stream == "framehizi" or selected_values_for_stream == "iframe"):
                        stream_value_table.append(selected_values_for_stream)                       
                    elif (selected_values_for_stream == "all"):
                        stream_value_table=['resolution','encoding','bitratetype','maxbitrate','framehizi','iframe','end']            
                    elif(selected_values_for_stream == "end"):
                         print("stream degerleri alindi.")            
                    else: 
                        print ("Deger tanimlanamadi.")           
            value_table.append(selected_values)        
        elif(selected_values == "end"):
            print("Degerler alindi.")
            value_table.append(selected_values)
        elif(selected_values == "get" or selected_values == "put"):
            print('1')        
        else:
            print("Deger tanimlanamadi ")
file='dokuman.xlsx'
os.path.isfile(file) # excel dosyasını kontrol et
wb=load_workbook(file)
ws=wb.active
for cell in ws['B']:
    ip_list.append(str(cell.value))
if value_table and PutorGet == "get":
    for i in range(2,len(ip_list)):
        ip = ip_list[i]        
        ping = os.system("ping -n 1 " + ip.strip())
        if (ping==0):
            print("#####################################")
            print("kamera ip :  "+ip.strip())
            for p in range(0,len(value_table)-1):
                print_value=value_table[p]
                print(print_value)
                XML=''
                if (print_value=='multicast') or (print_value=='mainstream'):
                    m='101'
                    value=19
                elif (print_value=='multicast2') or (print_value=='secondstream'):
                    m='102'
                    value=25
                elif (print_value=='multicast3') or (print_value=='thirdstream'):
                    m='103'
                    value=31
                
                PutorGet='get'
                cam=camera_connection(ip,user,password)  
                if (print_value == 'OnlineUserNumber' or print_value =='bitrate' ): #
                    if (print_value=='OnlineUserNumber'):
                        NumberofOnlineUser=requests.get('http://'+ip+'/ISAPI/System/workingstatus?format=json',auth=HTTPDigestAuth(user,password))
                        WorkingStatusLinkNum=re.findall("linkNum\":.*[0-9]+",NumberofOnlineUser.text)
                        WorkingStatusLinkNum1=re.sub('linkNum":\t', '',WorkingStatusLinkNum[0])
                        ws.cell(column=11, row=i+1).value=int(WorkingStatusLinkNum1) #Çekilen Stream Sayısı
                        print("Cekilen Stream Sayısı:      "+WorkingStatusLinkNum1)
                        wb.save("dokuman.xlsx")
                    elif (print_value=='bitrate'):
                        bitrate=requests.get('http://'+ip+'/ISAPI/System/workingstatus?format=json',auth=HTTPDigestAuth(user,password))
                        bitrate=re.findall("bitRate\":.*[0-9]+",bitrate.text)
                        bitrate1=re.sub('bitRate":\t', '',bitrate[0])
                        ws.cell(column=12, row=i+1).value=int(bitrate1) #bitrate
                        print("Toplam Bitrate: "+bitrate1)
                        wb.save("dokuman.xlsx") 
                else:
                    XML= XMLformat(cam)
                    if (print_value=='qos'):
                        print_value_on_xml='priorityValue'
                    elif (print_value=='multicast'):
                        print_value_on_xml='destIPAddress'
                    elif (print_value=='multicast2'):
                        print_value_on_xml='destIPAddress'
                    elif (print_value=='multicast3'):
                        print_value_on_xml='destIPAddress'
                    elif (print_value=='version'):
                        print_value_on_xml='firmwareVersion'
                    elif (print_value=='ntp'):
                        print_value_on_xml='ipAddress'
                    elif (print_value=='upnp') or (print_value=='motionDetection') or (print_value=='privacyMask') or (print_value=='ssh'):
                        print_value_on_xml='enabled'  
                    else:
                        print_value_on_xml=print_value

                    PrintDegeri1="<"+str(print_value_on_xml)+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"+"</"+str(print_value_on_xml)+">"
                    PrintDegeri2="<"+str(print_value_on_xml)+">"
                    PrintDegeri3="</"+str(print_value_on_xml)+">" 

                    if (print_value=='mainstream') or (print_value=='secondstream') or (print_value=='thirdstream'):
                        print("****"+print_value.upper()+"****")                        
                        for k in range(0,len(stream_value_table)-1):
                            stream_value=stream_value_table[k]
                            print(stream_value)
                            if (stream_value=='resolution'):
                                first_print_value_on_xml='videoResolutionWidth'
                                second_print_value_on_xml='videoResolutionHeight'
                            elif (stream_value=='encoding'):
                                first_print_value_on_xml='videoCodecType'
                            elif (stream_value=='bitratetype'):
                                first_print_value_on_xml='videoQualityControlType' 
                            elif (stream_value=='maxbitrate'):
                                first_print_value_on_xml='constantBitRate' 
                            elif (stream_value=='framehizi'):
                                first_print_value_on_xml='maxFrameRate'
                            elif (stream_value=='iframe'):
                                first_print_value_on_xml='GovLength' 
                            else:
                                first_print_value_on_xml=print_value

                            if (stream_value=='resolution'):
                                PrintDegeri1="<"+first_print_value_on_xml+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"
                                PrintDegeri2="<"+first_print_value_on_xml+">"
                                PrintDegeri3="</"+first_print_value_on_xml+">"
                                PrintDegeri4="<"+second_print_value_on_xml+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"
                                PrintDegeri5="<"+second_print_value_on_xml+">"
                                PrintDegeri6="</"+second_print_value_on_xml+">"
                                PrintDeger10 = re.findall(PrintDegeri1,XML)
                                PrintDeger11 = re.sub(PrintDegeri2, '',PrintDeger10[0])
                                PrintDeger11 = re.sub(PrintDegeri3, '',PrintDeger11)
                                PrintDeger12 = re.findall(PrintDegeri4,XML)
                                PrintDeger13 = re.sub(PrintDegeri5, '',PrintDeger12[0])
                                PrintDeger14 = PrintDeger11+"*"+PrintDeger13
                            else :
                                PrintDegeri1="<"+first_print_value_on_xml+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"
                                PrintDegeri2="<"+first_print_value_on_xml+">"
                                PrintDegeri3="</"+first_print_value_on_xml+">"
                                PrintDeger = re.findall(PrintDegeri1,XML)
                                PrintDeger1 = re.sub(PrintDegeri2, '',PrintDeger[0])

                            if (stream_value=='resolution'):
                                ws.cell(column=value, row=i+1).value = PrintDeger14
                                wb.save("dokuman.xlsx")
                            elif (stream_value=='bitratetype'):
                                ws.cell(column=value+1, row=i+1).value =PrintDeger1
                                wb.save("dokuman.xlsx")  
                            elif (stream_value=='maxbitrate'):
                                ws.cell(column=value+2, row=i+1).value =PrintDeger1
                                wb.save("dokuman.xlsx")
                            elif (stream_value=='framehizi'):
                                ws.cell(column=value+3, row=i+1).value =PrintDeger1
                                wb.save("dokuman.xlsx")            
                            elif (stream_value=='encoding'):
                                ws.cell(column=value+4, row=i+1).value = PrintDeger1
                                wb.save("dokuman.xlsx")                    
                            elif (stream_value=='iframe'):
                                ws.cell(column=value+5, row=i+1).value =PrintDeger1
                                wb.save("dokuman.xlsx")
                            else:
                                print("Stream Degerleri Excel'e yazilamadi")                            
                                
                    if (print_value=='qos'): 
                        PrintDeger = re.findall(PrintDegeri1,XML)
                        PrintDeger1=re.sub(PrintDegeri2,'',PrintDeger[0])
                        PrintDeger1=re.sub(PrintDegeri3,'',PrintDeger1)
                        PrintDeger3 = re.sub(PrintDegeri2, '',PrintDeger[1])
                        PrintDeger3 = re.sub(PrintDegeri3, '',PrintDeger3)
                        PrintDeger5 = re.sub(PrintDegeri2, '',PrintDeger[0]) 
                        PrintDeger5 = re.sub(PrintDegeri3, '',PrintDeger5) 
                        for k in range(0,len(qos_value_table)-1):
                            qos_value=qos_value_table[k]
                            if (qos_value=='videosesdscp'):
                                ws.cell(column=8, row=i+1).value=PrintDeger1
                                print("Video/Ses DSCP"+": "+PrintDeger1) 
                                wb.save("dokuman.xlsx")
                            elif (qos_value=='olayalarmdscp'):
                                ws.cell(column=9, row=i+1).value
                                print("Olay/Alarm DSCP"+": "+PrintDeger3)
                                wb.save("dokuman.xlsx")
                            elif (qos_value=='dscpyonetimi'):
                                ws.cell(column=10, row=i+1).value=PrintDeger5
                                print("DSCP Yönetimi"+": "+PrintDeger5)
                                wb.save("dokuman.xlsx")
                    elif ((print_value=='OnlineUserNumber') or (print_value=='bitrate')):
                        print('')                        
                    else :
                        PrintDeger = re.findall(PrintDegeri1,XML)
                        PrintDeger1 = re.sub(PrintDegeri2, '',PrintDeger[0])
                        PrintDeger1 = re.sub(PrintDegeri3, '',PrintDeger1)                    
                    if (print_value=='deviceName'):
                        ws.cell(column=1, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='version'):
                        ws.cell(column=14, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='multicast'):
                        ws.cell(column=3, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='multicast2'):
                        ws.cell(column=4, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='multicast3'):
                        ws.cell(column=5, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='speed'):
                       ws.cell(column=6, row=i+1).value=PrintDeger1
                       wb.save("dokuman.xlsx")
                    elif (print_value=='duplex'):
                        ws.cell(column=7, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='ntp'):
                        ws.cell(column=15, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='MTU'):
                        ws.cell(column=17, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='upnp'):
                        ws.cell(column=18, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='motionDetection'):
                        ws.cell(column=37, row=i+1).value=PrintDeger1
                        wb.save("dokuman.xlsx")
                    elif (print_value=='ssh'):
                        ws.cell(column=16, row=i+1).value=PrintDeger1
                    if not ((print_value=='qos') or (print_value=='mainstream') or (print_value=='secondstream') or (print_value=='thirdstream') or (print_value=='OnlineUserNumber') or (print_value=='bitrate')):
                        print(print_value+": "+PrintDeger1)
                        wb.save("dokuman.xlsx") 
                    
        else:
            print("#######################################################################")
            print(ip.strip()+" numarali kameranin baglantisi yok")
    wb.save("dokuman.xlsx")  
    wb.close()

elif value_table and PutorGet == "put":
    for i in range(2,len(ip_list)):
        ip = ip_list[i]
        ping = os.system("ping -n 1 " + ip.strip())
        if (ping==0):
            for p in range(0,len(value_table)-1):
                print_value=value_table[p]
                print(print_value)
                if (print_value=='multicast') or (print_value=='mainstream'):
                    m='101'
                    value=19
                elif (print_value=='multicast2') or (print_value=='secondstream'):
                    m='102'
                    value=25
                elif (print_value=='multicast3') or (print_value=='thirdstream'):
                    m='103'
                    value=31
                 
                if (print_value == 'OnlineUserNumber' or print_value =='bitrate' or print_value=='version' or print_value=='motionDetection'):
                    print('')
                else:                    
                    PutorGet = 'get'
                    cam=camera_connection(ip,user,password)
                    XML= XMLformat(cam)
                    PutorGet = 'put'

                if (print_value=='qos'):
                    print_value_on_cam='priorityValue'
                elif (print_value=='multicast'):
                    print_value_on_cam='destIPAddress'
                elif (print_value=='multicast2'):
                    print_value_on_cam='destIPAddress'
                elif (print_value=='multicast3'):
                    print_value_on_cam='destIPAddress'
                elif (print_value=='ntp'):
                    print_value_on_cam='ipAddress'
                elif (print_value=='upnp') or (print_value=='ssh'):
                    print_value_on_cam='enabled' 
                else:
                    print_value_on_cam=print_value

                PrintDegeri1="<"+print_value_on_cam+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"
                PrintDegeri2="<"+print_value_on_cam+">"
                PrintDegeri3="</"+print_value_on_cam+">"   

                if (print_value=='speed'):
                    put_value = ws.cell(column=6, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)                     
                    put_result=cam.System.Network.interfaces['1'].link(method='put',data=xml_value)
                elif (print_value=='duplex'):
                    put_value = ws.cell(column=7, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result=cam.System.Network.interfaces['1'].link(method='put',data=xml_value)
                elif (print_value=='MTU'):
                    put_value = ws.cell(column=17, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result=cam.System.Network.interfaces['1'].link(method='put',data=xml_value)
                elif (print_value=='upnp'):
                    put_value = ws.cell(column=18, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result=cam.System.Network.interfaces['1'].discovery(method='put',data=xml_value)
                elif (print_value=='qos'):
                    DSCPYonetimi = ws.cell(column=8, row=i+1).value
                    OlayAlarmDSCP = ws.cell(column=9, row=i+1).value
                    VideoSesDSCP = ws.cell(column=10, row=i+1).value
                    xml_value='<?xml version="1.0" encoding="UTF-8"?><DSCPList version="2.0" xmlns="http://www.hikvision.com/ver20/XMLSchema"><DSCP><id>1</id><enabled>TRUE</enabled><priorityValue>'+DSCPYonetimi+'</priorityValue><trafficType>devicemanagement</trafficType></DSCP><DSCP><id>2</id><enabled>TRUE</enabled><priorityValue>'+OlayAlarmDSCP+'</priorityValue><trafficType>commandcontrol</trafficType></DSCP><DSCP><id>3</id><enabled>TRUE</enabled><priorityValue>'+VideoSesDSCP+'</priorityValue><trafficType>video</trafficType></DSCP></DSCPList>'
                    put_result=cam.System.Network.qos.dscp(method='put',data=xml_value)
                    print(ip+": "+print_value+": Video/Ses DSCP: "+VideoSesDSCP+" Olay/Alarm DSCP: "+OlayAlarmDSCP+" DSCP Yonetimi: "+DSCPYonetimi)
                elif (print_value=='multicast'):
                    put_value = ws.cell(column=3, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result = cam.Streaming.channels[m](method='put', data=xml_value)
                elif (print_value=='multicast2'):
                    put_value = ws.cell(column=4, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result = cam.Streaming.channels[m](method='put', data=xml_value)
                elif (print_value=='multicast3'):
                    put_value = ws.cell(column=5, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result = cam.Streaming.channels[m](method='put', data=xml_value)
                elif (print_value=='deviceName'):
                    put_value = ws.cell(column=1, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result=cam.System.deviceInfo(method='put', data=xml_value)
                elif (print_value=='ntp'):
                    put_value = ws.cell(column=15, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result = cam.System.time.ntpservers(method='put',data=xml_value)
                elif (print_value=='privacyMask' and privacyMaskEnable != ''):
                    xml_value='<?xml version="1.0" encoding="UTF-8"?><PrivacyMask version="2.0" xmlns="http://www.hikvision.com/ver20/XMLSchema"><enabled>'+privacyMaskEnable+'</enabled><normalizedScreenSize><normalizedScreenWidth>704</normalizedScreenWidth><normalizedScreenHeight>576</normalizedScreenHeight></normalizedScreenSize><PrivacyMaskRegionList size="8"><PrivacyMaskRegion version="2.0" xmlns="http://www.hikvision.com/ver20/XMLSchema"><id>1</id><enabled>true</enabled><RegionCoordinatesList><RegionCoordinates><positionX>0</positionX><positionY>1</positionY></RegionCoordinates><RegionCoordinates><positionX>702</positionX><positionY>1</positionY></RegionCoordinates><RegionCoordinates><positionX>702</positionX><positionY>576</positionY></RegionCoordinates><RegionCoordinates><positionX>0</positionX><positionY>576</positionY></RegionCoordinates></RegionCoordinatesList></PrivacyMaskRegion></PrivacyMaskRegionList></PrivacyMask>'
                    put_result = cam.System.video.inputs.channels['1'].privacyMask(method='put',data=xml_value)
                elif (print_value=='ssh'):
                    put_value = ws.cell(column=16, row=i+1).value
                    xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                    put_result = cam.System.network.ssh(method='put',data=xml_value)
                elif (print_value=='mainstream') or (print_value=='secondstream') or (print_value=='thirdstream'):
                    print("****"+print_value.upper()+"****")
                    for c in range(0,len(stream_value_table)-1):
                        stream_value=stream_value_table[c]
                        if (stream_value=='encoding'):
                            print_value_on_cam='videoCodecType'
                        elif (stream_value=='bitratetype'):
                            print_value_on_cam='videoQualityControlType' 
                        elif (stream_value=='maxbitrate'):
                            print_value_on_cam='constantBitRate'   
                        elif (stream_value=='framehizi'):
                            print_value_on_cam='maxFrameRate' 
                        elif (stream_value=='iframe'):
                            print_value_on_cam='GovLength'
                        else:
                            print_value_on_cam = stream_value
                        if (stream_value=='resolution'):
                            videoResolutionWidth='videoResolutionWidth' 
                            videoResolutionHeight='videoResolutionHeight'
                            PrintDegeri1="<"+videoResolutionWidth+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"
                            PrintDegeri2="<"+videoResolutionWidth+">"
                            PrintDegeri3="</"+videoResolutionWidth+">"
                            PrintDegeri4="<"+videoResolutionHeight+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"
                            PrintDegeri5="<"+videoResolutionHeight+">"
                            PrintDegeri6="</"+videoResolutionHeight+">"
                        else: 
                            PrintDegeri1="<"+print_value_on_cam+">*[A-Za-z]*[0-9.]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*[-]*[A-Za-z]*[0-9]*[ ]*"
                            PrintDegeri2="<"+print_value_on_cam+">"
                            PrintDegeri3="</"+print_value_on_cam+">"

                        if (stream_value=='resolution'):
                            put_value = ws.cell(column=value, row=i+1).value
                            put_value_videoResolution = put_value.split("*")
                            yeniXML=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value_videoResolution[0]+PrintDegeri3, XML)
                            xml_value=re.sub(PrintDegeri4+PrintDegeri6,PrintDegeri5+put_value_videoResolution[1]+PrintDegeri6, yeniXML)
                            putsonuc=cam.Streaming.channels[m](method='put',present='text',data=xml_value)
                        elif (stream_value=='bitratetype'):
                            put_value = ws.cell(column=value+1, row=i+1).value
                            xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                            put_result=cam.Streaming.channels[m](method='put',present='text',data=xml_value)                            
                        elif (stream_value=='maxbitrate'):
                            put_value = ws.cell(column=value+2, row=i+1).value
                            xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                            put_result=cam.Streaming.channels[m](method='put',present='text',data=xml_value)                               
                        elif (stream_value=='framehizi'):
                            put_value = ws.cell(column=value+3, row=i+1).value
                            xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                            put_result=cam.Streaming.channels[m](method='put',present='text',data=xml_value)
                        elif (stream_value=='encoding'):
                            put_value = ws.cell(column=value+4, row=i+1).value
                            xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                            put_result=cam.Streaming.channels[m](method='put',present='text',data=xml_value)          
                        elif (stream_value=='iframe'):
                            put_value = ws.cell(column=value+5, row=i+1).value
                            xml_value=re.sub(PrintDegeri1+PrintDegeri3,PrintDegeri2+put_value+PrintDegeri3, XML)
                            put_result=cam.Streaming.channels[m](method='put',present='text',data=xml_value)          
                        
                        print(ip+": "+stream_value+": "+put_value)    
                else:
                    print('put komutunda sorun mevcut')
        else:
            print(ip.strip()+" numarali kameranin baglantisi yok ############################")

else:
    print('Hatalı put ya da get girdisi')
## frame hizi için 25 ise 2500 yazmalısın
